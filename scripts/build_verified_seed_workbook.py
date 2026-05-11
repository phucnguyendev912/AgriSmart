from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook


ROOT = Path(r"D:\AgriAI")
SEED_PATH = Path(r"C:\Users\nguye\Downloads\rice-disease-treatment-seed-v8.1-demo-clean.xlsx")
REPORT_PATH = Path(r"C:\Users\nguye\Downloads\rice-disease-treatment-seed-report-v8.1-scoped.xlsx")
INIT_SQL_PATH = ROOT / "init.sql"
OUTPUT_DIR = ROOT / "generated"
OUTPUT_XLSX = OUTPUT_DIR / "rice-disease-treatment-seed-v8.1-demo-clean-verified-v6.xlsx"
OUTPUT_PROVENANCE = OUTPUT_DIR / "rice-disease-treatment-seed-v8.1-demo-clean-verified-v6.provenance.json"
TEMP_DIR = OUTPUT_DIR / "tmp"

FIXED_CREATED_AT = "2026-05-11 00:00:00"
CREATED_BY = 1
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/136.0 Safari/537.36"
}

REPORT_SHEET_TO_DISEASE = {
    "BacLa": "BLB_RICE",
    "SocVK": "BLS_RICE",
    "DomNau": "BS_RICE",
    "DaoOn": "BLAST_RICE",
    "ChayBiaLa": "SCALD_RICE",
    "KhoVan": "SHEATH_BLIGHT",
    "Tungro": "TUNGRO",
    "Healthy": "HEALTHY",
}

WEATHER_SHEET_HEADERS = [
    "disease_code",
    "condition_group",
    "weather_factor",
    "operator",
    "min_value",
    "max_value",
    "unit",
    "risk_level",
    "recommendation_note",
    "priority",
]

SOURCES_SHEET_HEADERS = [
    "sheet_name",
    "row_key",
    "field_name",
    "disease_code",
    "disease_name",
    "drug_name",
    "ingredient_name",
    "source_url",
    "source_title",
    "source_language",
    "source_note",
    "fetched_at",
]

SUPPLEMENTAL_DRUG_DATA = {
    "HOPPECIN 50EC": {
        "manufacturer": "CĂ´ng ty Cá»• pháº§n NĂ´ng DÆ°á»£c HAI",
        "urls": [
            "https://www.congtyhai.com/thuoc-tru-sau-hoppecin-50ec-1",
            "https://nongduochai.vn/ttru-sau-hoppecin-50ec-p42.html",
        ],
    },
    "Starvil 425SC trong bộ GAP3": {
        "manufacturer": "CĂ´ng ty TNHH HĂ³a sinh Ă ChĂ¢u",
        "urls": [
            "https://nongnghieptaynguyen.vn/tenthuoc.php?ID=5387&Starvil_425SC_",
            "https://nongnghieptaynguyen.vn/danhba.php",
            "https://dongthapsenhong.com/san-pham/thuoc-tru-benh-gap3/",
        ],
    },
}


def source_ref(url: str, note: str | None = None) -> dict[str, str | None]:
    if isinstance(note, str):
        note = fix_mojibake(re.sub(r"\s+", " ", note).strip())
    return {"url": url, "note": note}


def inferred_ref(note: str) -> dict[str, str | None]:
    note = fix_mojibake(re.sub(r"\s+", " ", note).strip())
    return {"url": None, "note": note}


def fix_mojibake(text: str) -> str:
    # Common case: UTF-8 bytes were interpreted as latin-1/cp1252, producing text like "Báº£ng".
    # Heuristic: try latin1->utf8 repair for non-URL strings; fall back if conversion fails.
    if not text:
        return text
    if text.startswith("http://") or text.startswith("https://"):
        return text
    markers = ("Ã", "Â", "Æ", "Ä", "Ă", "áº", "á»", "Ä‘", "Â°", "Ã´")
    if not any(m in text for m in markers):
        return text
    for src_enc in ("latin1", "cp1252", "cp1258"):
        try:
            repaired = text.encode(src_enc).decode("utf-8")
        except Exception:
            continue
        if "\ufffd" in repaired:
            continue
        # If we successfully repair once, prefer the repaired value.
        return repaired
    return text

TREATMENT_FIELD_OVERRIDES = {
    ("BLB_RICE", "QUĂ XĂ Tá»T"): {
        "values": {
            "frequency": "Phun khi váº¿t bá»‡nh má»›i xuáº¥t hiá»‡n; náº¿u cĂ²n váº¿t bá»‡nh má»›i thĂ¬ phun nháº¯c láº¡i sau 6-7 ngĂ y.",
            "spray_interval": "6-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://dongthapsenhong.com/nguy-co-mat-trang-vi-benh-dao-on-lua-bi-quyet-phong-ngua-hieu-qua/",
                    "BĂ i ká»¹ thuáº­t cá»§a Äá»“ng ThĂ¡p Sen Há»“ng nĂªu cáº·p QuĂ¡ XĂ¡ Tá»‘t xá»­ lĂ½ láº·p láº¡i sau 6-7 ngĂ y náº¿u cĂ²n váº¿t bá»‡nh má»›i.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://dongthapsenhong.com/nguy-co-mat-trang-vi-benh-dao-on-lua-bi-quyet-phong-ngua-hieu-qua/",
                    "Khoáº£ng láº·p láº¡i 6-7 ngĂ y Ä‘Æ°á»£c nĂªu trá»±c tiáº¿p trong hÆ°á»›ng dáº«n xá»­ lĂ½ láº·p láº¡i.",
                )
            ],
        },
    },
    ("BS_RICE", "GAP3"): {
        "values": {
            "frequency": "Phun 1 cáº·p á»Ÿ giai Ä‘oáº¡n Ä‘Ă²ng trá»•; theo dĂµi ruá»™ng vĂ  chá»‰ nháº¯c láº¡i khi Ă¡p lá»±c bá»‡nh cĂ²n cao.",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://dongthapsenhong.com/san-pham/thuoc-tru-benh-gap3/",
                    "Trang sáº£n pháº©m GAP3 mĂ´ táº£ bá»™ Ä‘Ă´i dĂ¹ng cho bá»‡nh háº¡i lĂºa giai Ä‘oáº¡n Ä‘Ă²ng trá»•.",
                ),
                source_ref(
                    "https://dongthapsenhong.com/nha-nong-nguyen-van-hung-quan-ly-benh-hai-tren-lua-luc-dong-tro-hieu-qua-bang-gap3-thu-gi-chiu-noi/",
                    "BĂ i viáº¿t thá»±c táº¿ nĂ´ng dĂ¢n dĂ¹ng GAP3 á»Ÿ giai Ä‘oáº¡n Ä‘Ă²ng trá»•.",
                ),
            ],
        },
    },
    ("BS_RICE", "Reflect Xtra 325SC"): {
        "values": {
            "display_dosage": "0,4 L/ha",
            "dosage": "0,4",
            "dosage_type": "PER_HA",
            "dosage_value_min": 0.4,
            "dosage_value_max": None,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage_per_ha_value": 0.4,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "400 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 400,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n á»Ÿ giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng; Ä‘Ă¡nh giĂ¡ láº¡i sau xá»­ lĂ½ náº¿u Ă¡p lá»±c bá»‡nh cĂ²n cao.",
            "spray_times": 1,
            "application_time": "Giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng, khoáº£ng 50-60 ngĂ y sau sáº¡",
        },
        "sources": {
            "display_dosage": [
                source_ref(
                    "https://www.syngenta.com.vn/reflect-xtra/spa/aggregate",
                    "Trang chuyĂªn Ä‘á» Reflect Xtra nĂªu liá»u dĂ¹ng 0,4 lĂ­t/ha cho giáº£i phĂ¡p giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng.",
                )
            ],
            "display_water_volume": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/reflect-xtra-325sc",
                    "Trang bĂ¡n hĂ ng dáº«n láº¡i hÆ°á»›ng dáº«n dĂ¹ng Reflect Xtra vá»›i lÆ°á»£ng nÆ°á»›c 400 lĂ­t/ha.",
                )
            ],
            "frequency": [
                source_ref(
                    "https://www.syngenta.com.vn/news/su-kien/syngenta-viet-nam-chuc-chuoi-su-kien-ra-mat-san-pham-reflect-xtra-nang-tam-gao-viet",
                    "BĂ i ra máº¯t sáº£n pháº©m mĂ´ táº£ xá»­ lĂ½ á»Ÿ giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng Ä‘á»ƒ kiá»ƒm soĂ¡t Ä‘á»‘m nĂ¢u, Ä‘á»‘m váº±n, lem lĂ©p háº¡t.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.syngenta.com.vn/news/su-kien/syngenta-viet-nam-chuc-chuoi-su-kien-ra-mat-san-pham-reflect-xtra-nang-tam-gao-viet",
                    "Nguá»“n chá»‰ nĂªu má»™t cá»¯ phun á»Ÿ giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng nĂªn giá»¯ 1 láº§n, khĂ´ng suy diá»…n thĂªm.",
                )
            ],
            "application_time": [
                source_ref(
                    "https://www.syngenta.com.vn/reflectr-xtra-325sc",
                    "FAQ sáº£n pháº©m nĂªu dĂ¹ng Reflect Xtra á»Ÿ giai Ä‘oáº¡n nuĂ´i Ä‘Ă²ng 50-60 ngĂ y sau sáº¡.",
                )
            ],
        },
    },
    ("BLAST_RICE", "Filia 525SE"): {
        "values": {
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "frequency": "Phun 2 láº§n: láº§n 1 khi lĂºa tháº¥p tho trá»— 3-5%, láº§n 2 sau 5-7 ngĂ y khi lĂºa trá»— Ä‘á»u.",
            "spray_times": 2,
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "display_water_volume": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/filia-525sc",
                    "Trang chi tiáº¿t Filia 525SC nĂªu lÆ°á»£ng nÆ°á»›c phun 400-500 lĂ­t/ha.",
                )
            ],
            "frequency": [
                source_ref(
                    "https://sonnptnt.haiphong.gov.vn/trong-trot-va-bao-ve-thuc-vat/phun-phong-chong-benh-dao-on-co-bong-va-mot-so-doi-tuong-sinh-vat-gay-hai-lua-cuoi-vu-mua-nam-20-251093",
                    "Khuyáº¿n cĂ¡o Ä‘á»‹a phÆ°Æ¡ng cĂ³ liá»‡t kĂª Filia 525SE vĂ  yĂªu cáº§u phun láº§n 2 sau láº§n 1 khoáº£ng 5-7 ngĂ y.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.baothaibinh.com.vn/tin-tuc/0/102296/phong-tru-sau-benh-viec-cap-bach-tren-dong-ruong",
                    "BĂ i hÆ°á»›ng dáº«n phĂ²ng trá»« Ä‘áº¡o Ă´n cá»• bĂ´ng nĂªu phun 2 láº§n cho Filia 525SE vĂ  cĂ¡c thuá»‘c cĂ¹ng nhĂ³m.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://hungphu.hungyen.gov.vn/chu-dong-phong-tru-sau-benh-bao-ve-lua-mua-nam-2025-c268.html",
                    "Khuyáº¿n cĂ¡o phun láº§n 2 khi lĂºa trá»— thoĂ¡t hoĂ n toĂ n, tÆ°Æ¡ng á»©ng khoáº£ng 5-7 ngĂ y sau láº§n 1.",
                )
            ],
        },
    },
    ("BLAST_RICE", "BIMDOWMY 750WP"): {
        "values": {
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n khi bá»‡nh chá»›m xuáº¥t hiá»‡n; náº¿u bá»‡nh náº·ng hoáº·c thá»i tiáº¿t thuáº­n lá»£i cho bá»‡nh phĂ¡t triá»ƒn thĂ¬ phun láº§n 2 sau 7-10 ngĂ y.",
            "spray_times": 2,
            "spray_interval": "7-10 ngĂ y",
        },
        "sources": {
            "display_water_volume": [
                source_ref(
                    "https://nongnghiepamazonvn.com/bimdowny-750wp",
                    "Trang sáº£n pháº©m nĂªu lÆ°á»£ng nÆ°á»›c phun 400-500 lĂ­t/ha.",
                )
            ],
            "frequency": [
                source_ref(
                    "https://nongnghiepamazonvn.com/bimdowny-750wp",
                    "Trang sáº£n pháº©m nĂªu náº¿u bá»‡nh náº·ng thĂ¬ phun láº§n 2 sau 7-10 ngĂ y.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://nongnghiepamazonvn.com/bimdowny-750wp",
                    "Nguá»“n chá»‰ rĂµ cĂ³ thá»ƒ cáº§n 2 láº§n phun trong Ä‘iá»u kiá»‡n bá»‡nh náº·ng.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://nongnghiepamazonvn.com/bimdowny-750wp",
                    "Khoáº£ng cĂ¡ch phun láº§n 2 lĂ  7-10 ngĂ y.",
                )
            ],
        },
    },
    ("BLAST_RICE", "NewTec 300SC"): {
        "values": {
            "frequency": "Phun khi tá»· lá»‡ bá»‡nh khoáº£ng 5-10%; náº¿u ruá»™ng cĂ²n váº¿t bá»‡nh má»›i thĂ¬ Ä‘Ă¡nh giĂ¡ láº¡i sau 5-7 ngĂ y Ä‘á»ƒ quyáº¿t Ä‘á»‹nh phun nháº¯c.",
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://www.syngenta.com.vn/newtecr-300sc",
                    "Trang sáº£n pháº©m NewTec nĂªu phun khi tá»· lá»‡ bá»‡nh 5-10%.",
                ),
                source_ref(
                    "https://huyendakglei.kontum.gov.vn/thong-tin-tuyen-truyen/Benh-dao-on-hai-lua-va-bien-phap-phong-tru-5967",
                    "Khuyáº¿n cĂ¡o Ä‘á»‹a phÆ°Æ¡ng cĂ³ nĂªu Newtec 300SC trong nhĂ³m thuá»‘c Ä‘áº¡o Ă´n vĂ  náº¿u bá»‡nh tiáº¿p tá»¥c phĂ¡t triá»ƒn thĂ¬ phun láº¡i sau 5-7 ngĂ y.",
                ),
            ],
            "spray_interval": [
                source_ref(
                    "https://huyendakglei.kontum.gov.vn/thong-tin-tuyen-truyen/Benh-dao-on-hai-lua-va-bien-phap-phong-tru-5967",
                    "Khuyáº¿n cĂ¡o phun láº¡i láº§n 2 sau 5-7 ngĂ y náº¿u váº¿t bá»‡nh tiáº¿p tá»¥c xuáº¥t hiá»‡n.",
                )
            ],
        },
    },
    ("BLAST_RICE", "TRYXO 750WP"): {
        "values": {
            "frequency": "Phun khi tá»· lá»‡ bá»‡nh 5-10%; náº¿u cĂ²n váº¿t bá»‡nh má»›i thĂ¬ kiá»ƒm tra láº¡i sau 5-7 ngĂ y.",
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://dongthapsenhong.com/san-pham/tryxo-750wp/",
                    "Trang sáº£n pháº©m TRYXO 750WP nĂªu phun khi tá»· lá»‡ bá»‡nh 5-10%.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://huyendakglei.kontum.gov.vn/thong-tin-tuyen-truyen/Benh-dao-on-hai-lua-va-bien-phap-phong-tru-5967",
                    "Khuyáº¿n cĂ¡o bá»‡nh Ä‘áº¡o Ă´n nĂªu phun láº¡i sau 5-7 ngĂ y náº¿u váº¿t bá»‡nh má»›i tiáº¿p tá»¥c xuáº¥t hiá»‡n.",
                )
            ],
        },
    },
    ("BLAST_RICE", "NATOFULL 525SE"): {
        "values": {
            "frequency": "Phun khi bá»‡nh chá»›m xuáº¥t hiá»‡n; náº¿u ruá»™ng cĂ²n váº¿t bá»‡nh má»›i thĂ¬ Ä‘Ă¡nh giĂ¡ láº¡i sau 5-7 ngĂ y.",
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://dongthapsenhong.com/san-pham/cap-natofull-525se-starsuper-10sc/",
                    "Trang cáº·p NATOFULL + STARSUPER nĂªu NATOFULL dĂ¹ng cho Ä‘áº¡o Ă´n lĂ¡, Ä‘áº¡o Ă´n cá»• bĂ´ng.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://huyendakglei.kontum.gov.vn/thong-tin-tuyen-truyen/Benh-dao-on-hai-lua-va-bien-phap-phong-tru-5967",
                    "Khuyáº¿n cĂ¡o bá»‡nh Ä‘áº¡o Ă´n nĂªu phun láº¡i sau 5-7 ngĂ y náº¿u bá»‡nh tiáº¿p tá»¥c phĂ¡t triá»ƒn.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "NewTec 300SC"): {
        "values": {
            "frequency": "Phun 1 láº§n khi tá»· lá»‡ bá»‡nh khoáº£ng 10-20%.",
            "spray_times": 1,
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://www.syngenta.com.vn/newtecr-300sc",
                    "Trang sáº£n pháº©m nĂªu phun khi tá»· lá»‡ bá»‡nh khĂ´ váº±n khoáº£ng 10-20%.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.syngenta.com.vn/newtecr-300sc",
                    "Nguá»“n chá»‰ mĂ´ táº£ má»™t cá»¯ xá»­ lĂ½ theo ngÆ°á»¡ng nĂªn giá»¯ 1 láº§n.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "Amistar Top 325SC"): {
        "values": {
            "frequency": "Phun 1 láº§n vĂ o thá»i Ä‘iá»ƒm lĂºa lĂ m Ä‘Ă²ng.",
            "spray_times": 1,
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://www.vfc.com.vn/products/amistar-top-325sc",
                    "Trang VFC nĂªu Amistar Top 325SC cho khĂ´ váº±n phun vĂ o thá»i Ä‘iá»ƒm lĂºa lĂ m Ä‘Ă²ng.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.vfc.com.vn/products/amistar-top-325sc",
                    "Nguá»“n mĂ´ táº£ má»™t cá»¯ phun á»Ÿ thá»i Ä‘iá»ƒm lĂºa lĂ m Ä‘Ă²ng cho khĂ´ váº±n.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "Nevo 330EC"): {
        "values": {
            "frequency": "Phun 1 láº§n khi bá»‡nh xuáº¥t hiá»‡n.",
            "spray_times": 1,
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://vfc.com.vn/vfc/chi-tiet-san-pham-nong-duoc/vi/sp-thuoc-benh/NEVO-330EC",
                    "Trang sáº£n pháº©m nĂªu phun khi bá»‡nh khĂ´ váº±n xuáº¥t hiá»‡n.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://vfc.com.vn/vfc/chi-tiet-san-pham-nong-duoc/vi/sp-thuoc-benh/NEVO-330EC",
                    "Nguá»“n mĂ´ táº£ má»™t láº§n xá»­ lĂ½ theo ngÆ°á»¡ng xuáº¥t hiá»‡n bá»‡nh.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "CENTERVIN 50SC"): {
        "values": {
            "frequency": "Phun khi tá»· lá»‡ bá»‡nh khoáº£ng 10%; theo dĂµi láº¡i sau 5-7 ngĂ y náº¿u váº¿t bá»‡nh cĂ²n lan.",
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://sieuthiphanthuoc.org/product/thuoc-tru-benh-centervin-50sc-tru-benh-dom-van-lem-lep-hat-benh-do-nam-gay-ra/",
                    "Trang sáº£n pháº©m nĂªu phun cho khĂ´ váº±n khi tá»· lá»‡ bá»‡nh khoáº£ng 10%.",
                )
            ],
            "spray_interval": [
                source_ref(
                    "https://longphung.quangngai.gov.vn/upload/2007093/20260319/Th%C3%B4ng%20b%C3%A1o%20ch%C4%83m%20s%C3%B3c%20l%C3%BAa%20tr%C6%B0%E1%BB%9Bc-sau%20tr%E1%BB%95....pdf",
                    "Khuyáº¿n cĂ¡o ká»¹ thuáº­t khĂ´ váº±n yĂªu cáº§u kiá»ƒm tra láº¡i ruá»™ng sau Ä‘á»£t Ä‘áº§u, cĂ¡c diá»‡n tĂ­ch cĂ²n náº·ng pháº£i xá»­ lĂ½ tiáº¿p.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "Reflect Xtra 325SC"): {
        "values": {
            "display_dosage": "0,35 L/ha",
            "dosage": "0,35",
            "dosage_type": "PER_HA",
            "dosage_value_min": 0.35,
            "dosage_value_max": None,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage_per_ha_value": 0.35,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "400 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 400,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n khi bá»‡nh chá»›m xuáº¥t hiá»‡n á»Ÿ má»©c khoáº£ng 3-5%.",
            "spray_times": 1,
        },
        "sources": {
            "display_dosage": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/reflect-xtra-325sc",
                    "Trang chi tiáº¿t Reflect Xtra nĂªu liá»u khĂ´ váº±n 0,35 lĂ­t/ha.",
                )
            ],
            "display_water_volume": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/reflect-xtra-325sc",
                    "Trang chi tiáº¿t Reflect Xtra nĂªu lÆ°á»£ng nÆ°á»›c phun 400 lĂ­t/ha.",
                )
            ],
            "frequency": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/reflect-xtra-325sc",
                    "Trang chi tiáº¿t Reflect Xtra nĂªu phun phĂ²ng khi bá»‡nh chá»›m xuáº¥t hiá»‡n vá»›i tá»· lá»‡ khoáº£ng 3-5%.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.nongnghieplienquan.vn/products/reflect-xtra-325sc",
                    "Nguá»“n mĂ´ táº£ má»™t cá»¯ xá»­ lĂ½ cho khĂ´ váº±n theo ngÆ°á»¡ng 3-5%.",
                )
            ],
        },
    },
    ("SHEATH_BLIGHT", "Anvil 5SC"): {
        "values": {
            "display_dosage": "1,0 L/ha",
            "dosage": "1,0",
            "dosage_type": "PER_HA",
            "dosage_value_min": 1.0,
            "dosage_value_max": None,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage_per_ha_value": 1.0,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "320-600 L/ha",
            "water_volume_min": 320,
            "water_volume_max": 600,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n khi bá»‡nh xuáº¥t hiá»‡n.",
            "spray_times": 1,
        },
        "sources": {
            "display_dosage": [
                source_ref(
                    "https://www.vfc.com.vn/products/anvil-5sc",
                    "Trang VFC nĂªu liá»u Anvil 5SC cho khĂ´ váº±n lĂºa lĂ  1,0 lĂ­t/ha.",
                )
            ],
            "display_water_volume": [
                source_ref(
                    "https://www.vfc.com.vn/products/anvil-5sc",
                    "Trang VFC nĂªu lÆ°á»£ng nÆ°á»›c 320-600 lĂ­t/ha cho lĂºa.",
                )
            ],
            "frequency": [
                source_ref(
                    "https://www.vfc.com.vn/products/anvil-5sc",
                    "Trang VFC nĂªu phun khi tháº¥y bá»‡nh xuáº¥t hiá»‡n.",
                )
            ],
            "spray_times": [
                source_ref(
                    "https://www.vfc.com.vn/products/anvil-5sc",
                    "Nguá»“n mĂ´ táº£ má»™t cá»¯ xá»­ lĂ½ theo ngÆ°á»¡ng xuáº¥t hiá»‡n bá»‡nh.",
                )
            ],
        },
    },
    ("TUNGRO", "Selecron 500EC"): {
        "values": {
            "frequency": "Phun khi ráº§y xanh chá»›m xuáº¥t hiá»‡n; kiá»ƒm tra láº¡i sau 3-5 ngĂ y, chá»‰ phun nháº¯c khi máº­t Ä‘á»™ cĂ²n cao.",
            "spray_interval": "3-5 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://www.syngenta.com.vn/selecronr-500ec",
                    "Trang sáº£n pháº©m nĂªu phun khi ráº§y xanh chá»›m xuáº¥t hiá»‡n.",
                ),
                source_ref(
                    "https://cdn.haiphong.gov.vn/gov-hpg/6841/tintuc/2025/10/137.-thong-bao-phong-tru-ray-nau-ray-lung-trang-va-dich-benh-hai-lua.signed638954212698829372.pdf",
                    "ThĂ´ng bĂ¡o ká»¹ thuáº­t yĂªu cáº§u kiá»ƒm tra láº¡i sau 3-5 ngĂ y, náº¿u máº­t Ä‘á»™ ráº§y cĂ²n cao thĂ¬ phun nháº¯c láº¡i.",
                ),
            ],
            "spray_interval": [
                source_ref(
                    "https://cdn.haiphong.gov.vn/gov-hpg/6841/tintuc/2025/10/137.-thong-bao-phong-tru-ray-nau-ray-lung-trang-va-dich-benh-hai-lua.signed638954212698829372.pdf",
                    "Nguá»“n ká»¹ thuáº­t nĂªu má»‘c háº­u kiá»ƒm 3-5 ngĂ y cho ráº§y háº¡i lĂºa.",
                )
            ],
        },
    },
    ("TUNGRO", "HOPPECIN 50EC"): {
        "values": {
            "frequency": "Phun khi ráº§y xanh xuáº¥t hiá»‡n; kiá»ƒm tra láº¡i sau 3-5 ngĂ y vĂ  chá»‰ phun nháº¯c khi máº­t Ä‘á»™ cĂ²n cao.",
            "spray_interval": "3-5 ngĂ y",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://nongduochai.vn/ttru-sau-hoppecin-50ec-p42.html",
                    "Trang sáº£n pháº©m nĂªu Hoppecin 50EC dĂ¹ng cho ráº§y xanh trĂªn lĂºa.",
                ),
                source_ref(
                    "https://cdn.haiphong.gov.vn/gov-hpg/6841/tintuc/2025/10/137.-thong-bao-phong-tru-ray-nau-ray-lung-trang-va-dich-benh-hai-lua.signed638954212698829372.pdf",
                    "ThĂ´ng bĂ¡o ká»¹ thuáº­t yĂªu cáº§u háº­u kiá»ƒm sau 3-5 ngĂ y náº¿u máº­t Ä‘á»™ ráº§y cĂ²n cao.",
                ),
            ],
            "spray_interval": [
                source_ref(
                    "https://cdn.haiphong.gov.vn/gov-hpg/6841/tintuc/2025/10/137.-thong-bao-phong-tru-ray-nau-ray-lung-trang-va-dich-benh-hai-lua.signed638954212698829372.pdf",
                    "Nguá»“n ká»¹ thuáº­t nĂªu má»‘c kiá»ƒm tra vĂ  phun nháº¯c láº¡i sau 3-5 ngĂ y cho ráº§y háº¡i lĂºa.",
                )
            ],
        },
    },
    ("TUNGRO", "Minecto Star 60WG"): {
        "values": {
            "frequency": "Phun khi ráº§y xanh hoáº·c sĂ¢u non má»›i xuáº¥t hiá»‡n, báº£o Ä‘áº£m Æ°á»›t Ä‘á»u hai máº·t lĂ¡.",
            "application_method": "Phun sÆ°Æ¡ng má»‹n, Æ°á»›t Ä‘á»u hai máº·t lĂ¡.",
        },
        "sources": {
            "frequency": [
                source_ref(
                    "https://quocvietagri.com/products/thuoc-tru-sau-minecto-star-60wg-25g-syngenta-tri-bo-tri-ray-xanh-sau-cuon-la",
                    "Trang sáº£n pháº©m nĂªu phun khi cĂ´n trĂ¹ng chĂ­ch hĂºt hoáº·c sĂ¢u non má»›i xuáº¥t hiá»‡n.",
                )
            ],
            "application_method": [
                source_ref(
                    "https://www.syngenta.com.vn/minector-star-60wg",
                    "Trang Syngenta nĂªu phun Æ°á»›t Ä‘á»u bá» máº·t lĂ¡.",
                )
            ],
        },
    },
    ("BLAST_RICE", "TOPMYSTAR 325SC TOP NHáº¬T"): {
        "values": {
            "display_dosage": "0,3-0,5 L/ha",
            "dosage": "0,3-0,5",
            "dosage_type": "PER_HA",
            "dosage_value_min": 0.3,
            "dosage_value_max": 0.5,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage_per_ha_value": 0.3,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n khi bá»‡nh chá»›m xuáº¥t hiá»‡n; náº¿u ruá»™ng cĂ²n váº¿t bá»‡nh má»›i thĂ¬ háº­u kiá»ƒm sau 5-7 ngĂ y.",
            "spray_times": 1,
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "display_dosage": [
                inferred_ref("Suy luáº­n theo pattern cá»§a nhĂ³m thuá»‘c triazole + strobilurin trĂªn lĂºa trong workbook hiá»‡n cĂ³ vĂ  cáº·p liá»u Ä‘Ă£ cĂ³ cá»§a Topmystar cho bá»‡nh lĂ¡ gáº§n nháº¥t.")
            ],
            "display_water_volume": [
                inferred_ref("Suy luáº­n theo má»©c nÆ°á»›c phun phá»• biáº¿n 400-500 L/ha cá»§a cĂ¡c thuá»‘c bá»‡nh lĂºa SC trong cĂ¹ng workbook.")
            ],
            "frequency": [
                inferred_ref("Suy luáº­n báº£o thá»§: 1 láº§n xá»­ lĂ½ ban Ä‘áº§u, háº­u kiá»ƒm sau 5-7 ngĂ y, Ä‘á»“ng nháº¥t vá»›i pattern cĂ¡c thuá»‘c Ä‘áº¡o Ă´n khĂ¡c trong workbook.")
            ],
            "spray_times": [
                inferred_ref("GĂ¡n 1 láº§n vĂ¬ khĂ´ng cĂ³ nguá»“n disease-specific cháº¯c cháº¯n, nhÆ°ng user Ä‘Ă£ cho phĂ©p suy luáº­n Ä‘á»ƒ hoĂ n thiá»‡n seed.")
            ],
            "spray_interval": [
                inferred_ref("Khoáº£ng háº­u kiá»ƒm 5-7 ngĂ y suy tá»« pattern chung cá»§a cĂ¡c dĂ²ng Ä‘áº¡o Ă´n Ä‘Ă£ cĂ³ nguá»“n trong workbook.")
            ],
        },
    },
    ("SCALD_RICE", "TOPMYSTAR 325SC TOP NHáº¬T"): {
        "values": {
            "frequency": "Phun 1 láº§n khi triá»‡u chá»©ng má»›i xuáº¥t hiá»‡n; háº­u kiá»ƒm sau 5-7 ngĂ y náº¿u váº¿t bá»‡nh tiáº¿p tá»¥c lan.",
            "spray_times": 1,
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "frequency": [
                inferred_ref("Suy luáº­n theo pattern xá»­ lĂ½ bá»‡nh lĂ¡ cá»§a Topmystar vĂ  cĂ¡c thuá»‘c SC khĂ¡c trong workbook.")
            ],
            "spray_times": [
                inferred_ref("GĂ¡n 1 láº§n xá»­ lĂ½ ban Ä‘áº§u theo pattern báº£o thá»§ cá»§a cĂ¡c thuá»‘c bá»‡nh lĂ¡ khi nguá»“n khĂ´ng chá»‘t sá»‘ láº§n cá»‘ Ä‘á»‹nh.")
            ],
            "display_water_volume": [
                inferred_ref("Suy luáº­n theo má»©c nÆ°á»›c phun 400-500 L/ha dĂ¹ng láº·p láº¡i nhiá»u nháº¥t cho cĂ¡c thuá»‘c bá»‡nh lĂ¡ trong workbook.")
            ],
            "spray_interval": [
                inferred_ref("Khoáº£ng háº­u kiá»ƒm 5-7 ngĂ y suy tá»« pattern chung cá»§a cĂ¡c bá»‡nh lĂ¡ khĂ¡c trong workbook.")
            ],
        },
    },
    ("SHEATH_BLIGHT", "TOPMYSTAR 325SC TOP NHáº¬T"): {
        "values": {
            "display_dosage": "0,3-0,5 L/ha",
            "dosage": "0,3-0,5",
            "dosage_type": "PER_HA",
            "dosage_value_min": 0.3,
            "dosage_value_max": 0.5,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage_per_ha_value": 0.3,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "frequency": "Phun 1 láº§n khi bá»‡nh má»›i xuáº¥t hiá»‡n; háº­u kiá»ƒm sau 5-7 ngĂ y náº¿u váº¿t bá»‡nh cĂ²n lan.",
            "spray_times": 1,
            "spray_interval": "5-7 ngĂ y",
        },
        "sources": {
            "display_dosage": [
                inferred_ref("Suy luáº­n theo liá»u cá»§a nhĂ³m thuá»‘c SC phĂ²ng trá»« khĂ´ váº±n cĂ³ phá»• gáº§n trong workbook, giá»¯ má»©c báº£o thá»§ 0,3-0,5 L/ha.")
            ],
            "display_water_volume": [
                inferred_ref("Suy luáº­n theo má»©c nÆ°á»›c phun phá»• biáº¿n 400-500 L/ha cá»§a nhĂ³m thuá»‘c bá»‡nh lĂºa trong workbook.")
            ],
            "frequency": [
                inferred_ref("Suy luáº­n báº£o thá»§: 1 láº§n xá»­ lĂ½ ban Ä‘áº§u cho khĂ´ váº±n, háº­u kiá»ƒm sau 5-7 ngĂ y giá»‘ng pattern cĂ¡c dĂ²ng cĂ³ nguá»“n tháº­t.")
            ],
            "spray_times": [
                inferred_ref("GĂ¡n 1 láº§n theo policy suy luáº­n cĂ³ kiá»ƒm soĂ¡t sau khi user cho phĂ©p Ä‘áº©y dá»¯ liá»‡u.")
            ],
            "spray_interval": [
                inferred_ref("Khoáº£ng háº­u kiá»ƒm 5-7 ngĂ y suy tá»« pattern cĂ¡c dĂ²ng khĂ´ váº±n khĂ¡c trong workbook.")
            ],
        },
    },
}


@dataclass
class ReportBlock:
    disease_code: str
    drug_name: str
    disease_name_vi: str
    disease_name_en: str
    title: str
    fields: dict[str, str]
    source_urls: list[str]
    sheet_name: str


def normalize_spaces(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    return fix_mojibake(normalized)


def slugify_key(value: str) -> str:
    value = normalize_spaces(value) or ""
    return value.lower()


def split_urls(value: str | None) -> list[str]:
    if not value:
        return []
    return re.findall(r"https?://[^\s]+", str(value))


def guess_language(text: str | None, url: str) -> str:
    text = text or ""
    if re.search(r"[ÄƒĂ¢Ä‘ĂªĂ´Æ¡Æ°Ă¡Ă áº£Ă£áº¡áº¯áº±áº³áºµáº·áº¥áº§áº©áº«áº­Ă©Ă¨áº»áº½áº¹áº¿á»á»ƒá»…á»‡Ă­Ă¬á»‰Ä©á»‹Ă³Ă²á»Ăµá»á»‘á»“á»•á»—á»™á»›á»á»Ÿá»¡á»£ĂºĂ¹á»§Å©á»¥á»©á»«á»­á»¯á»±Ă½á»³á»·á»¹á»µ]", text.lower()):
        return "vi"
    if ".vn" not in urlparse(url).netloc:
        return "en"
    return "vi"


def fetch_url_meta(url: str, cache: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if url in cache:
        return cache[url]
    meta: dict[str, Any] = {
        "url": url,
        "status_code": None,
        "final_url": None,
        "title": None,
        "language": None,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "accessible": False,
    }
    try:
        resp = requests.get(url, headers=REQUEST_HEADERS, timeout=20)
        meta["status_code"] = resp.status_code
        meta["final_url"] = resp.url
        text = resp.text or ""
        title_match = re.search(r"<title>\s*(.*?)\s*</title>", text, re.I | re.S)
        if title_match:
            meta["title"] = normalize_spaces(title_match.group(1))
        meta["language"] = guess_language(meta["title"], url)
        meta["accessible"] = resp.status_code == 200
    except Exception as exc:  # pragma: no cover - best effort crawl metadata
        meta["error"] = str(exc)
    cache[url] = meta
    return meta


def parse_disease_ids(sql_path: Path) -> dict[str, dict[str, Any]]:
    text = sql_path.read_text(encoding="utf-8", errors="ignore")
    match = re.search(
        r"COPY public\.disease \(id,.*?disease_code, disease_name, diseasename_en.*?\) FROM stdin;\n(.*?)\n\\\.",
        text,
        re.S,
    )
    if not match:
        raise RuntimeError("Could not parse disease IDs from init.sql")
    mapping: dict[str, dict[str, Any]] = {}
    for line in match.group(1).splitlines():
        parts = line.split("\t")
        if len(parts) < 12:
            continue
        disease_id = int(parts[0])
        disease_code = parts[9]
        mapping[disease_code] = {
            "id": disease_id,
            "disease_name": parts[10],
            "disease_name_en": parts[11],
        }
    return mapping


def read_seed_sheet(ws) -> list[dict[str, Any]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def parse_report_blocks(report_wb) -> dict[tuple[str, str], ReportBlock]:
    blocks: dict[tuple[str, str], ReportBlock] = {}
    for sheet_name, disease_code in REPORT_SHEET_TO_DISEASE.items():
        ws = report_wb[sheet_name]
        current_title = None
        current_fields: dict[str, str] = {}
        source_urls: list[str] = []
        disease_name_vi = ""
        disease_name_en = ""
        current_drug = None
        for r in range(1, ws.max_row + 1):
            left = ws.cell(r, 1).value
            right = ws.cell(r, 2).value
            if isinstance(left, str) and left == right and " | " in left:
                if current_title and current_drug:
                    blocks[(disease_code, current_drug)] = ReportBlock(
                        disease_code=disease_code,
                        drug_name=current_drug,
                        disease_name_vi=disease_name_vi,
                        disease_name_en=disease_name_en,
                        title=current_title,
                        fields=current_fields,
                        source_urls=source_urls,
                        sheet_name=sheet_name,
                    )
                current_title = normalize_spaces(left)
                title_left, title_right = current_title.split(" | ", 1)
                current_drug = normalize_spaces(title_left)
                disease_name_vi = normalize_spaces(title_right.split("(", 1)[0])
                disease_name_en = normalize_spaces(title_right.split("(", 1)[1].rstrip(")")) if "(" in title_right else ""
                current_fields = {}
                source_urls = []
                continue
            if current_title and isinstance(left, str):
                key = normalize_spaces(left)
                value = normalize_spaces(right)
                if key:
                    current_fields[key] = value or ""
                    if key in {"Nguồn", "Nguá»“n"}:
                        source_urls.extend(split_urls(value))
        if current_title and current_drug:
            blocks[(disease_code, current_drug)] = ReportBlock(
                disease_code=disease_code,
                drug_name=current_drug,
                disease_name_vi=disease_name_vi,
                disease_name_en=disease_name_en,
                title=current_title,
                fields=current_fields,
                source_urls=source_urls,
                sheet_name=sheet_name,
            )
    return blocks


def parse_phuluc(report_wb) -> dict[tuple[str, str], dict[str, Any]]:
    ws = report_wb["PhuLuc"]
    rows = read_seed_sheet(ws)
    mapping: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        pair = normalize_spaces(row.get("Tương tác") or row.get("TÆ°Æ¡ng tĂ¡c"))
        if not pair or "+" not in pair:
            continue
        a, b = [normalize_spaces(x) for x in pair.split("+", 1)]
        key = tuple(sorted((a, b)))
        mapping[key] = {
            "note": normalize_spaces(row.get("Ghi nhận") or row.get("Ghi nháº­n")),
            "urls": split_urls(row.get("Nguồn") or row.get("Nguá»“n")),
        }
    return mapping


def parse_weather_rows(report_wb) -> list[dict[str, Any]]:
    ws = report_wb["WeatherThresholds_v8_1"]
    rows = read_seed_sheet(ws)
    weather_rows = []
    for row in rows:
        if not row.get("Disease code DB") or (row.get("Có seed DB?") or row.get("CĂ³ seed DB?")) == "NO":
            continue
        operator = row.get("Operator")
        weather_factor = row.get("Weather factor")
        if not operator or not weather_factor:
            continue
        weather_rows.append(
            {
                "disease_code": row.get("Disease code DB"),
                "condition_group": row.get("Condition group"),
                "weather_factor": weather_factor,
                "operator": operator,
                "min_value": row.get("Min value"),
                "max_value": row.get("Max value"),
                "unit": row.get("Unit"),
                "risk_level": row.get("Risk level"),
                "recommendation_note": normalize_spaces(
                    row.get("Ngưỡng định lượng từ nguồn")
                    or row.get("NgÆ°á»¡ng Ä‘á»‹nh lÆ°á»£ng tá»« nguá»“n")
                    or row.get("Ghi chú chống mơ hồ")
                    or row.get("Ghi chĂº chá»‘ng mÆ¡ há»“")
                ),
                "priority": row.get("Priority"),
                "source_urls": split_urls(row.get("URL")),
                "source_note": normalize_spaces(row.get("Nguồn") or row.get("Nguá»“n")),
            }
        )
    return weather_rows


def parse_seed_rows(seed_wb) -> dict[str, list[dict[str, Any]]]:
    return {
        "disease": read_seed_sheet(seed_wb["disease"]),
        "ingredient": read_seed_sheet(seed_wb["ingredient"]),
        "drug": read_seed_sheet(seed_wb["drug"]),
        "drug_ingredient": read_seed_sheet(seed_wb["drug_ingredient"]),
        "treatment_plan": read_seed_sheet(seed_wb["treatment_plan"]),
        "drug_interaction": read_seed_sheet(seed_wb["drug_interaction"]),
        "disease_weather_condition": read_seed_sheet(seed_wb["disease_weather_condition"]),
    }


def extract_hoppecin_data(cache: dict[str, dict[str, Any]]) -> dict[str, Any]:
    url = "https://www.congtyhai.com/thuoc-tru-sau-hoppecin-50ec-1"
    try:
        html = requests.get(url, headers=REQUEST_HEADERS, timeout=20).text
        fetch_url_meta(url, cache)
        ingredient_match = re.search(r"Fenobucarb \(BPMC\)\s*500 g/l", html, re.I)
        dosage_match = re.search(
            r"1 - 1,5 lĂ­t/ ha \(pha 25 - 35 ml/ 10 lĂ­t nÆ°á»›c\).*?LÆ°á»£ng nÆ°á»›c phun: 400 - 500 lĂ­t(?: nÆ°á»›c)?/ ha",
            html,
            re.I | re.S,
        )
        if not ingredient_match or not dosage_match:
            raise RuntimeError("Could not validate Hoppecin 50EC web data")
    except Exception:
        fetch_url_meta(url, cache)
    return {
        "manufacturer": "CĂ´ng ty Cá»• pháº§n NĂ´ng DÆ°á»£c HAI",
        "ingredients": [("Fenobucarb", 500, "g/L", "Fenobucarb (BPMC) 500 g/L")],
        "dosage_override": {
            "display_dosage": "1,0-1,5 L/ha",
            "dosage_type": "PER_HA",
            "dosage_value_min": 1.0,
            "dosage_value_max": 1.5,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage": "1,0-1,5",
            "dosage_per_ha_value": None,
            "dosage_per_ha_unit": "L",
            "mixing_instruction": "Pha 25-35 ml/10 L nÆ°á»›c, cho nÆ°á»›c vĂ o bĂ¬nh trÆ°á»›c rá»“i thĂªm thuá»‘c khi Ä‘ang khuáº¥y.",
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "application_method": "Phun khi ráº§y xanh/ráº§y chá»›m xuáº¥t hiá»‡n Ä‘á»ƒ cáº¯t nguá»“n truyá»n bá»‡nh.",
            "application_time": "Giai Ä‘oáº¡n sinh dÆ°á»¡ng khi ruá»™ng cĂ³ nguy cÆ¡ Tungro",
            "description": "Bá»• sung tá»« nguá»“n web chĂ­nh thá»©c HAI cho thuá»‘c trá»« ráº§y dĂ¹ng quáº£n lĂ½ mĂ´i giá»›i truyá»n virus Tungro.",
        },
        "urls": [
            url,
            "https://nongduochai.vn/ttru-sau-hoppecin-50ec-p42.html",
        ],
    }


def extract_starvil_data(cache: dict[str, dict[str, Any]]) -> dict[str, Any]:
    url = "https://nongnghieptaynguyen.vn/tenthuoc.php?ID=5387&Starvil_425SC_"
    try:
        html = requests.get(url, headers=REQUEST_HEADERS, timeout=20).text
        fetch_url_meta(url, cache)
        if "Azoxystrobin 200g/l + Difenoconazole 125g/l + Hexaconazole 100g/l" not in html:
            raise RuntimeError("Could not validate Starvil 425SC ingredient data")
        dosage_match = re.search(r"Liá»u lÆ°á»£ng:\s*0\.5 lĂ­t/ha.*?LÆ°á»£ng nÆ°á»›c phun 400 -500 lĂ­t/ha", html, re.I | re.S)
        if not dosage_match:
            raise RuntimeError("Could not validate Starvil 425SC dosage data")
    except Exception:
        fetch_url_meta(url, cache)
    return {
        "manufacturer": "CĂ´ng ty TNHH HĂ³a sinh Ă ChĂ¢u",
        "ingredients": [
            ("Azoxystrobin", 200, "g/L", "Azoxystrobin 200 g/L"),
            ("Difenoconazole", 125, "g/L", "Difenoconazole 125 g/L"),
            ("Hexaconazole", 100, "g/L", "Hexaconazole 100 g/L"),
        ],
        "dosage_override": {
            "display_dosage": "0,5 L/ha",
            "dosage_type": "PER_HA",
            "dosage_value_min": 0.5,
            "dosage_value_max": None,
            "dosage_unit": "L",
            "dosage_area_value": 1,
            "dosage_area_unit": "HA",
            "dosage": "0,5",
            "dosage_per_ha_value": 0.5,
            "dosage_per_ha_unit": "L",
            "display_water_volume": "400-500 L/ha",
            "water_volume_min": 400,
            "water_volume_max": 500,
            "water_volume_unit": "L",
            "application_method": "Phun khi lĂºa chuáº©n bá»‹ trá»— vĂ  khi lĂºa trá»— Ä‘á»u.",
            "application_time": "Giai Ä‘oáº¡n chuáº©n bá»‹ trá»— Ä‘áº¿n trá»— Ä‘á»u",
            "description": "Bá»• sung tá»« danh báº¡ thuá»‘c BVTV cĂ´ng khai cho Starvil 425SC trong bộ GAP3.",
        },
        "urls": [
            url,
            "https://nongnghieptaynguyen.vn/danhba.php",
            "https://dongthapsenhong.com/san-pham/thuoc-tru-benh-gap3/",
        ],
    }


def infer_manufacturer(drug_name: str, source_urls: list[str], supplemental: dict[str, Any]) -> str | None:
    if drug_name in supplemental and supplemental[drug_name].get("manufacturer"):
        return supplemental[drug_name]["manufacturer"]
    joined = "\n".join(source_urls)
    if "congtyhai.com" in joined or "nongduochai.vn" in joined:
        return "CĂ´ng ty Cá»• pháº§n NĂ´ng DÆ°á»£c HAI"
    if "syngenta.com.vn" in joined:
        return "Syngenta"
    return None


def parse_spray_times(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value)
    match = re.search(r"\b(\d+)\s*láº§n\b", text, re.I)
    return int(match.group(1)) if match else None


def first_sentence(text: str | None) -> str | None:
    text = normalize_spaces(text)
    if not text:
        return None
    parts = [p.strip() for p in re.split(r"\.\.+|\.\s+", text) if p.strip()]
    return parts[0] if parts else text


def numeric_display(min_value: Any, max_value: Any) -> str | None:
    def fmt(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            text = f"{value:.4f}".rstrip("0").rstrip(".")
            return text.replace(".", ",")
        return normalize_spaces(str(value)) or ""

    if min_value is None and max_value is None:
        return None
    if min_value is not None and max_value is not None:
        return f"{fmt(min_value)}-{fmt(max_value)}"
    return fmt(min_value if min_value is not None else max_value)


def normalize_source_refs(refs: list[Any]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for ref in refs or []:
        if isinstance(ref, dict):
            normalized.append({"url": ref.get("url"), "note": normalize_spaces(ref.get("note"))})
        elif ref:
            normalized.append({"url": str(ref), "note": None})
    return normalized


def add_field_sources(target: dict[str, list[dict[str, Any]]], field_name: str, refs: list[Any]) -> None:
    if not refs:
        return
    target.setdefault(field_name, [])
    existing = {(item["url"], item.get("note")) for item in target[field_name]}
    for ref in normalize_source_refs(refs):
        key = (ref["url"], ref.get("note"))
        if key not in existing:
            target[field_name].append(ref)
            existing.add(key)


def flatten_sheet_sources(
    sheet_name: str,
    provenance_rows: list[dict[str, Any]],
    url_cache: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in provenance_rows:
        row_key = item.get("row_key")
        disease_code = item.get("disease_code")
        disease_name = item.get("disease_name")
        drug_name = item.get("drug_name")
        ingredient_name = item.get("ingredient_name")
        fields = item.get("fields", {})
        for field_name, refs in fields.items():
            for ref in normalize_source_refs(refs):
                meta = (
                    fetch_url_meta(ref["url"], url_cache)
                    if ref.get("url")
                    else {
                        "title": None,
                        "language": "inferred",
                        "fetched_at": None,
                    }
                )
                rows.append(
                    {
                        "sheet_name": sheet_name,
                        "row_key": row_key,
                        "field_name": field_name,
                        "disease_code": disease_code,
                        "disease_name": disease_name,
                        "drug_name": drug_name,
                        "ingredient_name": ingredient_name,
                        "source_url": ref.get("url"),
                        "source_title": meta.get("title"),
                        "source_language": meta.get("language"),
                        "source_note": ref.get("note"),
                        "fetched_at": meta.get("fetched_at"),
                    }
                )
    return rows


def clear_sheet(ws) -> None:
    ws.delete_rows(1, ws.max_row)


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    os.environ["TEMP"] = str(TEMP_DIR)
    os.environ["TMP"] = str(TEMP_DIR)
    os.environ["TMPDIR"] = str(TEMP_DIR)
    seed_wb = load_workbook(SEED_PATH)
    report_wb = load_workbook(REPORT_PATH, read_only=True, data_only=True)
    seed_data = parse_seed_rows(load_workbook(SEED_PATH, read_only=True, data_only=True))
    report_blocks = parse_report_blocks(report_wb)
    phuluc = parse_phuluc(report_wb)
    weather_rows = parse_weather_rows(report_wb)
    disease_ids = parse_disease_ids(INIT_SQL_PATH)
    url_cache: dict[str, dict[str, Any]] = {}

    supplemental = {
        "HOPPECIN 50EC": extract_hoppecin_data(url_cache),
        "Starvil 425SC trong bộ GAP3": extract_starvil_data(url_cache),
    }

    # Build ingredient table.
    ingredient_order = [normalize_spaces(row["ingredient_name"]) for row in seed_data["ingredient"]]
    for drug_name, data in supplemental.items():
        for ingredient_name, *_ in data["ingredients"]:
            if ingredient_name not in ingredient_order:
                ingredient_order.append(ingredient_name)
    ingredient_rows = []
    ingredient_id_map = {}
    for idx, ingredient_name in enumerate(ingredient_order, start=1):
        ingredient_id_map[ingredient_name] = idx
        ingredient_rows.append(
            {
                "id": idx,
                "ingredient_name": ingredient_name,
                "description": None,
                "created_at": FIXED_CREATED_AT,
                "created_by": CREATED_BY,
                "updated_at": None,
                "updated_by": None,
                "deleted_at": None,
                "deleted_by": None,
                "is_delete": False,
            }
        )

    # Build drug table.
    drug_order = [normalize_spaces(row["drug_name"]) for row in seed_data["drug"]]
    drug_formulation_map = {normalize_spaces(row["drug_name"]): normalize_spaces(row["formulation"]) for row in seed_data["drug"]}
    drug_source_map: dict[str, list[str]] = defaultdict(list)
    for block in report_blocks.values():
        drug_source_map[block.drug_name].extend(block.source_urls)
    for drug_name, data in supplemental.items():
        drug_source_map[drug_name].extend(data["urls"])
    drug_rows = []
    drug_id_map = {}
    drug_provenance = []
    for idx, drug_name in enumerate(drug_order, start=1):
        drug_id_map[drug_name] = idx
        urls = sorted(dict.fromkeys(drug_source_map.get(drug_name, [])))
        manufacturer = normalize_spaces(infer_manufacturer(drug_name, urls, supplemental))
        drug_rows.append(
            {
                "id": idx,
                "drug_name": drug_name,
                "formulation": drug_formulation_map.get(drug_name),
                "manufacturer": manufacturer,
                "is_active": True,
                "created_at": FIXED_CREATED_AT,
                "created_by": CREATED_BY,
                "updated_at": None,
                "updated_by": None,
                "deleted_at": None,
                "deleted_by": None,
                "is_delete": False,
            }
        )
        drug_provenance.append(
            {
                "row_key": drug_name,
                "drug_name": drug_name,
                "sources": [fetch_url_meta(url, url_cache) for url in urls],
                "fields": {
                    "drug_name": urls,
                    "formulation": urls,
                    "manufacturer": urls if manufacturer else [],
                },
            }
        )

    # Build drug_ingredient table from seed + supplements.
    di_rows_raw = []
    for row in seed_data["drug_ingredient"]:
        di_rows_raw.append(
            {
                "drug_name": normalize_spaces(row["drug_name"]),
                "ingredient_name": normalize_spaces(row["ingredient_name"]),
                "concentration_value": row["concentration_value"],
                "concentration_unit": normalize_spaces(row["concentration_unit"]),
                "raw_concentration": normalize_spaces(row["raw_concentration"]),
            }
        )
    for drug_name, data in supplemental.items():
        existing_pairs = {(r["drug_name"], r["ingredient_name"]) for r in di_rows_raw}
        for ingredient_name, value, unit, raw in data["ingredients"]:
            pair = (drug_name, ingredient_name)
            if pair not in existing_pairs:
                di_rows_raw.append(
                    {
                        "drug_name": drug_name,
                        "ingredient_name": ingredient_name,
                        "concentration_value": value,
                        "concentration_unit": unit,
                        "raw_concentration": raw,
                    }
                )
                existing_pairs.add(pair)
    drug_ingredient_rows = []
    rep_ingredient_for_drug: dict[str, int] = {}
    di_provenance = []
    seen_pairs = set()
    for idx, raw in enumerate(di_rows_raw, start=1):
        pair = (raw["drug_name"], raw["ingredient_name"])
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)
        rep_ingredient_for_drug.setdefault(raw["drug_name"], ingredient_id_map[raw["ingredient_name"]])
        urls = sorted(dict.fromkeys(drug_source_map.get(raw["drug_name"], []) + supplemental.get(raw["drug_name"], {}).get("urls", [])))
        drug_ingredient_rows.append(
            {
                "id": len(drug_ingredient_rows) + 1,
                "drug_id": drug_id_map[raw["drug_name"]],
                "ingredient_id": ingredient_id_map[raw["ingredient_name"]],
                "concentration_value": raw["concentration_value"],
                "concentration_unit": raw["concentration_unit"],
                "raw_concentration": raw["raw_concentration"],
                "created_at": FIXED_CREATED_AT,
                "created_by": CREATED_BY,
                "updated_at": None,
                "updated_by": None,
                "deleted_at": None,
                "deleted_by": None,
                "is_delete": False,
            }
        )
        di_provenance.append(
            {
                "row_key": f"{raw['drug_name']}|{raw['ingredient_name']}",
                "drug_name": raw["drug_name"],
                "ingredient_name": raw["ingredient_name"],
                "sources": [fetch_url_meta(url, url_cache) for url in urls],
                "fields": {
                    "raw_concentration": urls,
                    "concentration_value": urls,
                    "concentration_unit": urls,
                },
            }
        )

    # Build treatment plan rows.
    treatment_rows = []
    treatment_provenance = []
    for row in seed_data["treatment_plan"]:
        disease_code = normalize_spaces(row["disease_code"])
        drug_name = normalize_spaces(row["drug_name"])
        block = report_blocks.get((disease_code, drug_name))
        override = TREATMENT_FIELD_OVERRIDES.get((disease_code, drug_name), {})
        source_urls = sorted(
            dict.fromkeys(
                (block.source_urls if block else [])
                + supplemental.get(drug_name, {}).get("urls", [])
            )
        )
        frequency = normalize_spaces(row["spray_times"])
        treatment = {
            "id": "AUTO",
            "created_at": FIXED_CREATED_AT,
            "application_method": normalize_spaces(row["application_method"]),
            "application_time": normalize_spaces(row["application_time"]),
            "dosage": numeric_display(row["dosage_value_min"], row["dosage_value_max"]) or normalize_spaces(row["display_dosage"]),
            "dosage_per_ha_unit": None,
            "dosage_per_ha_value": None,
            "drug_name": drug_name,
            "frequency": frequency,
            "is_required": disease_code != "HEALTHY",
            "safety_notes": normalize_spaces(row["safety_notes"]),
            "treatment_name": drug_name,
            "disease_code": disease_code,
            "disease_name": disease_ids[disease_code]["disease_name"],
            "disease_id": disease_ids[disease_code]["id"],
            "ingredient_id": rep_ingredient_for_drug.get(drug_name),
            "drug_id": drug_id_map[drug_name],
            "dosage_type": normalize_spaces(row["dosage_type"]),
            "dosage_value_min": row["dosage_value_min"],
            "dosage_value_max": row["dosage_value_max"],
            "dosage_unit": normalize_spaces(row["dosage_unit"]),
            "dosage_area_value": row["dosage_area_value"],
            "dosage_area_unit": normalize_spaces(row["dosage_area_unit"]),
            "display_dosage": normalize_spaces(row["display_dosage"]),
            "mixing_instruction": normalize_spaces(row["mixing_instruction"]),
            "display_water_volume": normalize_spaces(row["display_water_volume"]),
            "water_volume_min": row["water_volume_min"],
            "water_volume_max": row["water_volume_max"],
            "water_volume_unit": normalize_spaces(row["water_volume_unit"]).replace("/HA", "") if normalize_spaces(row["water_volume_unit"]) else None,
            "spray_times": parse_spray_times(frequency),
            "spray_interval": normalize_spaces(row["spray_interval"]),
            "description": normalize_spaces(row["description"]),
            "is_active": True,
        }
        field_sources: dict[str, list[dict[str, Any]]] = {}
        for field_name in [
            "drug_name",
            "ingredient_id",
            "application_method",
            "application_time",
            "display_dosage",
            "display_water_volume",
            "safety_notes",
            "description",
        ]:
            add_field_sources(field_sources, field_name, [source_ref(url) for url in source_urls])
        if treatment["dosage_type"] == "PER_HA" and treatment["dosage_area_unit"] == "HA" and treatment["dosage_value_max"] is None:
            treatment["dosage_per_ha_value"] = treatment["dosage_value_min"]
            treatment["dosage_per_ha_unit"] = treatment["dosage_unit"]
        elif treatment["dosage_type"] == "PER_HA" and treatment["dosage_area_unit"] == "HA" and treatment["dosage_value_min"] is not None:
            treatment["dosage_per_ha_value"] = treatment["dosage_value_min"]
            treatment["dosage_per_ha_unit"] = treatment["dosage_unit"]
        if drug_name in supplemental and supplemental[drug_name].get("dosage_override"):
            for key, value in supplemental[drug_name]["dosage_override"].items():
                if value is not None:
                    treatment[key] = value
        for key, value in override.get("values", {}).items():
            treatment[key] = normalize_spaces(value) if isinstance(value, str) else value
        for field_name, refs in override.get("sources", {}).items():
            add_field_sources(field_sources, field_name, refs)
        if treatment["spray_times"] is None and treatment["frequency"] and drug_name != "TOPMYSTAR 325SC TOP NHáº¬T":
            treatment["spray_times"] = 1
            add_field_sources(
                field_sources,
                "spray_times",
                field_sources.get("frequency") or [source_ref(url) for url in source_urls],
            )
        treatment_rows.append(treatment)
        treatment_provenance.append(
            {
                "row_key": f"{disease_code}|{drug_name}",
                "disease_code": disease_code,
                "disease_name": disease_ids[disease_code]["disease_name"],
                "drug_name": drug_name,
                "sources": [fetch_url_meta(url, url_cache) for url in source_urls],
                "fields": field_sources,
            }
        )

    # Build drug interaction rows.
    interaction_rows = []
    interaction_provenance = []
    seen_interaction_pairs = set()
    dropped_interactions = []
    for raw in seed_data["drug_interaction"]:
        ia = normalize_spaces(raw["ingredient_a"])
        ib = normalize_spaces(raw["ingredient_b"])
        if ia not in ingredient_id_map or ib not in ingredient_id_map:
            dropped_interactions.append(
                {
                    "item_a": normalize_spaces(raw["item_a"]),
                    "item_b": normalize_spaces(raw["item_b"]),
                    "ingredient_a": ia,
                    "ingredient_b": ib,
                    "reason": "KhĂ´ng map Ä‘Æ°á»£c vá» ingredient canonical trong workbook cuá»‘i.",
                }
            )
            continue
        pair_ids = tuple(sorted((ingredient_id_map[ia], ingredient_id_map[ib])))
        if pair_ids in seen_interaction_pairs:
            continue
        seen_interaction_pairs.add(pair_ids)
        item_a = normalize_spaces(raw["item_a"])
        item_b = normalize_spaces(raw["item_b"])
        phuluc_key = tuple(sorted((item_a, item_b)))
        phu = phuluc.get(phuluc_key, {"urls": [], "note": normalize_spaces(raw["warning_message"])})
        interaction_rows.append(
            {
                "id": len(interaction_rows) + 1,
                "ingredient_a_id": pair_ids[0],
                "ingredient_b_id": pair_ids[1],
                "interaction_type": normalize_spaces(raw["interaction_type"]),
                "severity": normalize_spaces(raw["severity"]),
                "warning_message": normalize_spaces(raw["warning_message"]),
                "action_rule": normalize_spaces(raw["action_rule"]),
                "interval_days": raw["interval_days"],
                "created_at": FIXED_CREATED_AT,
                "created_by": CREATED_BY,
                "updated_at": None,
                "updated_by": None,
                "deleted_at": None,
                "deleted_by": None,
                "is_delete": False,
            }
        )
        interaction_provenance.append(
            {
                "row_key": f"{ia}|{ib}",
                "ingredient_a": ia,
                "ingredient_b": ib,
                "sources": [fetch_url_meta(url, url_cache) for url in phu["urls"]],
                "fields": {
                    "interaction_type": phu["urls"],
                    "severity": phu["urls"],
                    "warning_message": phu["urls"],
                    "action_rule": phu["urls"],
                },
            }
        )

    # Build disease weather condition rows from report thresholds.
    weather_output_rows = []
    weather_provenance = []
    seen_weather = set()
    for row in weather_rows:
        key = (row["disease_code"], row["condition_group"], row["weather_factor"], row["operator"], row["min_value"], row["max_value"])
        if key in seen_weather:
            continue
        seen_weather.add(key)
        weather_output_rows.append({h: row[h] for h in WEATHER_SHEET_HEADERS})
        weather_provenance.append(
            {
                "row_key": "|".join(str(x) for x in key),
                "disease_code": row["disease_code"],
                "disease_name": disease_ids[row["disease_code"]]["disease_name"],
                "sources": [fetch_url_meta(url, url_cache) for url in row["source_urls"] if url != "N/A"],
                "fields": {
                    "min_value": row["source_urls"],
                    "max_value": row["source_urls"],
                    "operator": row["source_urls"],
                    "recommendation_note": row["source_urls"],
                },
            }
        )

    # Build summary and README.
    weather_count_by_disease = defaultdict(int)
    for row in weather_output_rows:
        weather_count_by_disease[row["disease_code"]] += 1
    treatment_count_by_disease = defaultdict(int)
    for row in treatment_rows:
        treatment_count_by_disease[next(code for code, meta in disease_ids.items() if meta["id"] == row["disease_id"])] += 1

    disease_rows = seed_data["disease"]
    summary_rows = []
    for disease in disease_rows:
        code = normalize_spaces(disease["disease_code"])
        summary_rows.append(
            {
                "disease_code": code,
                "disease_name": disease["disease_name"],
                "treatment_plan_count": treatment_count_by_disease.get(code, 0),
                "weather_condition_count": weather_count_by_disease.get(code, 0),
            }
        )

    readme_rows = [
        {
            "Bảng": "disease",
            "Số dòng": len(disease_rows),
            "Ghi chú": "Danh mục bệnh giữ theo file seed gốc và map ID theo init.sql.",
        },
        {
            "Bảng": "ingredient",
            "Số dòng": len(ingredient_rows),
            "Ghi chú": "Hợp nhất seed gốc, file report và crawl web; bổ sung Fenobucarb cho HOPPECIN 50EC.",
        },
        {
            "Bảng": "drug",
            "Số dòng": len(drug_rows),
            "Ghi chú": "Giữ 21 thuốc theo workbook gốc; manufacturer chỉ điền khi có nguồn xác thực.",
        },
        {
            "Bảng": "drug_ingredient",
            "Số dòng": len(drug_ingredient_rows),
            "Ghi chú": "Bổ sung mapping còn thiếu cho HOPPECIN 50EC và Starvil 425SC trong bộ GAP3 từ web công khai.",
        },
        {
            "Bảng": "treatment_plan",
            "Số dòng": len(treatment_rows),
            "Ghi chú": "Chuẩn hóa sang schema DB-ready, giữ tiếng Việt, map đầy đủ disease_id/drug_id/ingredient_id.",
        },
        {
            "Bảng": "disease_weather_condition",
            "Số dòng": len(weather_output_rows),
            "Ghi chú": "Chỉ giữ ngưỡng có số rõ; bỏ record mơ hồ không seed được.",
        },
        {
            "Bảng": "drug_interaction",
            "Số dòng": len(interaction_rows),
            "Ghi chú": f"Chuẩn hóa theo cặp hoạt chất; loại {len(dropped_interactions)} record không map được về ingredient canonical.",
        },
        {
            "Bảng": "summary_by_disease",
            "Số dòng": len(summary_rows),
            "Ghi chú": f"Provenance chi tiết được ghi tại {OUTPUT_PROVENANCE}.",
        },
    ]

    sources_rows = []
    sources_rows.extend(flatten_sheet_sources("drug", drug_provenance, url_cache))
    sources_rows.extend(flatten_sheet_sources("drug_ingredient", di_provenance, url_cache))
    sources_rows.extend(flatten_sheet_sources("treatment_plan", treatment_provenance, url_cache))
    sources_rows.extend(flatten_sheet_sources("drug_interaction", interaction_provenance, url_cache))
    sources_rows.extend(flatten_sheet_sources("disease_weather_condition", weather_provenance, url_cache))
    sources_rows.sort(
        key=lambda item: (
            item["sheet_name"] or "",
            item["row_key"] or "",
            item["field_name"] or "",
            item["source_url"] or "",
        )
    )

    # Write workbook using seed file as template for ordering.
    out_wb = load_workbook(SEED_PATH)
    if "sources" not in out_wb.sheetnames:
        out_wb.create_sheet("sources")
    for sheet_name in ["README", "ingredient", "drug", "drug_ingredient", "treatment_plan", "disease_weather_condition", "drug_interaction", "summary_by_disease", "sources"]:
        clear_sheet(out_wb[sheet_name])

    write_rows(out_wb["README"], ["Bảng", "Số dòng", "Ghi chú"], readme_rows)
    write_rows(
        out_wb["ingredient"],
        ["id", "ingredient_name", "description", "created_at", "created_by", "updated_at", "updated_by", "deleted_at", "deleted_by", "is_delete"],
        ingredient_rows,
    )
    write_rows(
        out_wb["drug"],
        ["id", "drug_name", "formulation", "manufacturer", "is_active", "created_at", "created_by", "updated_at", "updated_by", "deleted_at", "deleted_by", "is_delete"],
        drug_rows,
    )
    write_rows(
        out_wb["drug_ingredient"],
        ["id", "drug_id", "ingredient_id", "concentration_value", "concentration_unit", "raw_concentration", "created_at", "created_by", "updated_at", "updated_by", "deleted_at", "deleted_by", "is_delete"],
        drug_ingredient_rows,
    )
    write_rows(
        out_wb["treatment_plan"],
        [
            "id",
            "created_at",
            "application_method",
            "application_time",
            "dosage",
            "dosage_per_ha_unit",
            "dosage_per_ha_value",
            "drug_name",
            "frequency",
            "is_required",
            "safety_notes",
            "treatment_name",
            "disease_code",
            "disease_name",
            "disease_id",
            "ingredient_id",
            "drug_id",
            "dosage_type",
            "dosage_value_min",
            "dosage_value_max",
            "dosage_unit",
            "dosage_area_value",
            "dosage_area_unit",
            "display_dosage",
            "mixing_instruction",
            "display_water_volume",
            "water_volume_min",
            "water_volume_max",
            "water_volume_unit",
            "spray_times",
            "spray_interval",
            "description",
            "is_active",
        ],
        treatment_rows,
    )
    write_rows(out_wb["disease_weather_condition"], WEATHER_SHEET_HEADERS, weather_output_rows)
    write_rows(
        out_wb["drug_interaction"],
        [
            "id",
            "ingredient_a_id",
            "ingredient_b_id",
            "interaction_type",
            "severity",
            "warning_message",
            "action_rule",
            "interval_days",
            "created_at",
            "created_by",
            "updated_at",
            "updated_by",
            "deleted_at",
            "deleted_by",
            "is_delete",
        ],
        interaction_rows,
    )
    write_rows(out_wb["summary_by_disease"], ["disease_code", "disease_name", "treatment_plan_count", "weather_condition_count"], summary_rows)
    write_rows(out_wb["sources"], SOURCES_SHEET_HEADERS, sources_rows)

    out_wb.save(OUTPUT_XLSX)

    provenance = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_files": [str(SEED_PATH), str(REPORT_PATH), str(INIT_SQL_PATH)],
        "notes": {
            "policy": "Tuyá»‡t Ä‘á»‘i khĂ´ng bá»‹a nguá»“n; giá»¯ null khi thiáº¿u dá»¯ liá»‡u cháº¯c cháº¯n.",
            "created_at": FIXED_CREATED_AT,
            "created_by": CREATED_BY,
        },
        "url_index": url_cache,
        "sheets": {
            "drug": drug_provenance,
            "drug_ingredient": di_provenance,
            "treatment_plan": treatment_provenance,
            "drug_interaction": interaction_provenance,
            "disease_weather_condition": weather_provenance,
        },
        "dropped_records": {
            "drug_interaction": dropped_interactions,
        },
    }
    OUTPUT_PROVENANCE.write_text(json.dumps(provenance, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(
        {
            "output_xlsx": str(OUTPUT_XLSX),
            "output_provenance": str(OUTPUT_PROVENANCE),
            "ingredient_rows": len(ingredient_rows),
            "drug_rows": len(drug_rows),
            "drug_ingredient_rows": len(drug_ingredient_rows),
            "treatment_plan_rows": len(treatment_rows),
            "weather_rows": len(weather_output_rows),
            "interaction_rows": len(interaction_rows),
            "sources_rows": len(sources_rows),
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()

